import os
import time
import random
from playwright.sync_api import sync_playwright
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import csv

def launch_live_dynamic_engine(url, wait_selector=None):
    """
    Launches a live browser session and yields the active page controller.
    This keeps the browser open so you can interact with dynamic components live.
    """
    # 1. Start Playwright
    p = sync_playwright().start()
    
    # 2. Launch browser with a real User-Agent to bypass basic anti-bot blocks
    browser = p.chromium.launch(headless=False)
    context = browser.new_context(
        user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        viewport={"width": 1920, "height": 1080}
    )
    page = context.new_page()
    
    try:
        print(f"[Live Engine] Opening: {url}")
        # Wait until network activity settles
        page.goto(url, wait_until="networkidle", timeout=100000)
        
        if wait_selector:
            page.wait_for_selector(wait_selector, timeout=10000)
            
        # Return the LIVE page, the browser, and the playwright instance so we can close them later
        return page, browser, p
        
    except Exception as e:
        print(f"[Live Engine Error]: {e}")
        page.close()
        context.close()
        browser.close()
        p.stop()
        return None, None, None


def extract_safe_text(parent_locator, selector, default=None):
    """Safely extracts text from a sub-locator if it exists."""
    target = parent_locator.locator(selector)
    if target.count() > 0:
        return target.inner_text().strip()
    return default

def extract_safe_attribute(parent_locator, selector, attribute, default=None):
    # Check if selector is empty
    if selector == "":
        target = parent_locator
    else:
        target = parent_locator.locator(selector)
    
    # Now count() will work because target is the element itself, not a new query
    if target.count() > 0:
        return target.get_attribute(attribute)
    return default

def append_to_csv(file_path, data_dict, fieldnames=None):
    """
    Safely appends a single row of data (as a dictionary) to a CSV file.
    Automatically creates the file and writes the header if the file does not exist yet.
    """
    try:
        # 1. Determine if we need to write a header row
        file_exists = os.path.exists(file_path) and os.path.getsize(file_path) > 0
        
        # 2. If fieldnames aren't explicitly provided, use the dictionary keys
        if not fieldnames:
            fieldnames = list(data_dict.keys())
            
        # 3. Open the file in append ('a') mode with utf-8 encoding to prevent special character crashes
        with open(file_path, mode="a", newline="", encoding="utf-8") as csv_file:
            writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
            
            # If the file is brand new, write the top header row first
            if not file_exists:
                writer.writeheader()
                
            # 4. Write the actual data row
            writer.writerow(data_dict)
            
    except Exception as e: 
        print(f"Critical Storage Error writing to {file_path}: {e}")






def convert_csv_to_styled_excel(
    csv_path, 
    excel_path, 
    sheet_name="Data Export", 
    brand_color="2A4B7C"
):
    """
    Completely modularized Excel engine. Pass any CSV, any destination path, 
    any sheet name, and any hex color code to instantly match a client's branding.
    """
    try:
        # 1. Read raw data via pandas and pipe it to standard Excel
        df = pd.read_csv(csv_path)
        df.to_excel(excel_path, index=False, sheet_name=sheet_name)
        
        # 2. Open workbook with openpyxl to inject professional formatting
        wb = openpyxl.load_workbook(excel_path)
        ws = wb[sheet_name]
        
        # Ensure gridlines are visible
        ws.views.sheetView[0].showGridLines = True
        
        # Freeze the top header row
        ws.freeze_panes = "A2"
        
        # Dynamic brand styling mapping
        header_fill = PatternFill(start_color=brand_color, end_color=brand_color, fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        body_font = Font(name="Segoe UI", size=10, bold=False, color="333333")
        
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_alignment = Alignment(horizontal="left", vertical="center")
        right_alignment = Alignment(horizontal="right", vertical="center")
        
        thin_border_side = Side(border_style="thin", color="E0E0E0")
        body_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        # Style Header Row
        ws.row_dimensions[1].height = 28
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            
        # Style Data Rows
        for row in range(2, ws.max_row + 1):
            ws.row_dimensions[row].height = 20
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = body_font
                cell.border = body_border
                
                # Align numeric data right, text left
                if isinstance(cell.value, (int, float)):
                    cell.alignment = right_alignment
                    if "price" in str(ws.cell(1, col).value).lower():
                        cell.number_format = '"$"#,##0.00'
                else:
                    cell.alignment = left_alignment
                    
        # Auto-fit Column Widths cleanly
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        wb.save(excel_path)
        print(f"Success! Branded Excel workbook generated at: {excel_path}")
        
    except Exception as e:
        print(f"Excel Generation Error: {e}")



